#import pandas as pd

#import "C:\Users\daad2295\AppData\Local\Python\pythoncore-3.14-64\Lib\site-packages\openpyxl"
import openpyxl
#how to check PYTHONPATH? this program runs fine on my desktop but I can't run it in VS Code on my desktop?
#-------------------------------------------------------------------------------------------------------------------------------------------------------
#Define Student class
#-------------------------------------------------------------------------------------------------------------------------------------------------------
class Student:
    stu_info_keys = []
    stu_info_values = []
    stu_info_dict = {}
    plos_file_name = ""
    Slate_file = ""

    def __init__(self, plos_file_name="template",Slate_file="Plan of Study Templater.xlsx",stu_info_keys=["CU-SID","Person Name","Pronouns","Study Plan Codes","Subplan Code 1","Subplan Code 2","Admit Term", "Person Visa Type", "Transfer Courses","BAM Supplement","Student Classes", "Waivers", "Overall GPA", "Other Study Plans"]):
        self.Slate_file = Slate_file
        self.stu_info_keys = stu_info_keys
        self.plos_file_name=plos_file_name
#-------------------------------------------------------------------------------------------------------------------------------------------------------
#Function - get # of plans of study
#-------------------------------------------------------------------------------------------------------------------------------------------------------
def get_num_plos(num_plos):
    #access and open Slate file
    #path="C:\\Users\\daad2295\\Desktop\\PLOS_Maker\\Plan of Study Templater.xlsx"
    #path="R:\\gradadmin\\Grad Program\\3. Daniel Adams_GPS\\Student Forms\\Plan of Study\\PLOS_Maker\\Plan of Study Templater.xlsx"
    path='Plan of Study Templater.xlsx'
    slate_wb=openpyxl.load_workbook(path)
    slate_sheet=slate_wb.active

    #identify number of plos to create, minus the header row
    num_plos = slate_sheet.max_row - 1
    return num_plos
#-------------------------------------------------------------------------------------------------------------------------------------------------------
#Function - import data from Slate to PLOS.py
#-------------------------------------------------------------------------------------------------------------------------------------------------------
def import_from_slate(plos_list, row_num):
    #access and open Slate file
    #path="C:\\Users\\daad2295\\Desktop\\PLOS_Maker\\Plan of Study Templater.xlsx"
    #path="R:\\gradadmin\\Grad Program\\3. Daniel Adams_GPS\\Student Forms\\Plan of Study\\PLOS_Maker\\Plan of Study Templater.xlsx"
    path='Plan of Study Templater.xlsx'
    slate_wb=openpyxl.load_workbook(path)
    slate_sheet=slate_wb.active
        
    current = Student() #variables should be template at this point
    c=1
    r = row_num #user_provided
    current.stu_info_values.clear()
    for key in current.stu_info_keys:
        cell_obj=slate_sheet.cell(r,c)
        current.stu_info_values.append(cell_obj.value)
        c +=1
    current.stu_info_dict = dict(zip(current.stu_info_keys,current.stu_info_values))
    
    #update the path variable here, so we can differentiate between MSNE and MSCPS plos
    if ("CSEN-MSCPS" in current.stu_info_dict["Study Plan Codes"]):
        #current.plos_file_name="C:\\Users\\daad2295\\Desktop\\PLOS_Maker\\Plan of Study_MSCPS.xlsx"
        #current.plos_file_name="R:\\gradadmin\\Grad Program\\3. Daniel Adams_GPS\\Student Forms\\Plan of Study\\PLOS_Maker\\Plan of Study_MSCPS.xlsx"
        current.plos_file_name='Plan of Study_MSCPS.xlsx'
    elif("NTEN-MSNE" in current.stu_info_dict["Study Plan Codes"]):
        #current.plos_file_name="C:\\Users\\daad2295\\Desktop\\PLOS_Maker\\Plan of Study_MSNE.xlsx"
        #current.plos_file_name="R:\\gradadmin\\Grad Program\\3. Daniel Adams_GPS\\Student Forms\\Plan of Study\\PLOS_Maker\\Plan of Study_MSNE.xlsx"
        current.plos_file_name='Plan of Study_MSNE.xlsx'
    elif("AINT-MSAIP" in current.stu_info_dict["Study Plan Codes"]):
        current.plos_file_name='Plan of Study_MSAIP.xlsx'
    plos_list.append(current)

#-------------------------------------------------------------------------------------------------------------------------------------------------------
#actual program start 
#-------------------------------------------------------------------------------------------------------------------------------------------------------

plos_list = []
num_plos = 0
num_plos = get_num_plos(num_plos)
row_num = 1 

for i in range (num_plos):
    row_num += 1
    import_from_slate(plos_list, row_num)

for stu in plos_list:
    path = stu.plos_file_name
    plos_wb=openpyxl.load_workbook(path)
    plos_sheet=plos_wb['Data']

    #save as new file with new name
    #new_file_path = "C:\\Users\\daad2295\\Desktop\\PLOS_Maker\\"
    new_file_path="R:\\gradadmin\\Grad Program\\3. Daniel Adams_GPS\\Student Forms\\Plan of Study\\"

    if ("CSEN-MSCPS" in stu.stu_info_dict["Study Plan Codes"]):
        if (stu.stu_info_dict["Subplan Code 1"] != None):
            if(stu.stu_info_dict["Subplan Code 2"] != None):
                new_file_name = "Plos_" + stu.stu_info_dict["Person Name"] + "_MSCPS, " + stu.stu_info_dict["Subplan Code 1"] + ", " + stu.stu_info_dict["Subplan Code 2"] + ".xlsx"
            else:
                new_file_name = "Plos_" + stu.stu_info_dict["Person Name"] + "_MSCPS, " + stu.stu_info_dict["Subplan Code 1"] + ".xlsx"
        elif (stu.stu_info_dict["Subplan Code 2"] != None):
            new_file_name = "Plos_" + stu.stu_info_dict["Person Name"] + "_MSCPS-" + stu.stu_info_dict["Subplan Code 2"] + ".xlsx"
        else:
            new_file_name = "Plos_" + stu.stu_info_dict["Person Name"] + "_MSCPS.xlsx"
            
    elif ("NTEN-MSNE" in stu.stu_info_dict["Study Plan Codes"]):
        if (stu.stu_info_dict["Subplan Code 1"] != None):
            new_file_name = "Plos_" + stu.stu_info_dict["Person Name"] + "_MSNE, " + stu.stu_info_dict["Subplan Code 1"] + ".xlsx"
        elif (stu.stu_info_dict["Subplan Code 2"] != None):
            new_file_name = "Plos_" + stu.stu_info_dict["Person Name"] + "_MSNE, " + stu.stu_info_dict["Subplan Code 2"] + ".xlsx"
        else:
            new_file_name = "Plos_" + stu.stu_info_dict["Person Name"] + "_MSNE.xlsx"

    elif ("AINT-MSAIP" in stu.stu_info_dict["Study Plan Codes"]):
        if (stu.stu_info_dict["Subplan Code 1"] != None):
            new_file_name = "Plos_" + stu.stu_info_dict["Person Name"] + "_MSAIP, " + stu.stu_info_dict["Subplan Code 1"] + ".xlsx"
        elif (stu.stu_info_dict["Subplan Code 2"] != None):
            new_file_name = "Plos_" + stu.stu_info_dict["Person Name"] + "_MSAIP, " + stu.stu_info_dict["Subplan Code 2"] + ".xlsx"
        else:
            new_file_name = "Plos_" + stu.stu_info_dict["Person Name"] + "_MSAIP.xlsx"
    
    print(new_file_name)
    new_file_path += new_file_name
    plos_wb.save(new_file_path)

   #fill it with Student Information
    e = plos_sheet['D2']
    e.value = stu.stu_info_dict["CU-SID"]
    e = plos_sheet['D3']
    e.value = stu.stu_info_dict["Person Name"]
    e = plos_sheet['D4']
    e.value = stu.stu_info_dict["Pronouns"]
    e = plos_sheet['D5']
    e.value = stu.stu_info_dict["Study Plan Codes"]
    e = plos_sheet['D6']
    e.value = stu.stu_info_dict["Subplan Code 1"]
    e = plos_sheet['D7']
    e.value = stu.stu_info_dict["Subplan Code 2"]
    e = plos_sheet['D8']
    e.value = stu.stu_info_dict["Admit Term"]
    e = plos_sheet['D9']
    e.value = stu.stu_info_dict["Person Visa Type"]
    e = plos_sheet['D10']
    e.value = stu.stu_info_dict["Waivers"]
    e = plos_sheet['D11']
    e.value = stu.stu_info_dict["Overall GPA"]
    e = plos_sheet['D13']
    e.value = stu.stu_info_dict["Other Study Plans"]
        
    #process Student Classes column
    student_classes=[]
    temp = ""
    temp = stu.stu_info_dict["Student Classes"]
    if (temp != None):
        student_classes=temp.split('*')

    #process Transfer Classes column
    transfer_classes=[]
    temp = ""
    temp = stu.stu_info_dict["Transfer Courses"]
    if (temp != None):
        transfer_classes = temp.split('*')
        index=0
        for transfer_class in range(len(transfer_classes)):
            student_classes.append(transfer_classes[index])
            index+=1
        
    #process BAM Supplement column ONLY IF TRANSFER CLASSES WAS EMPTY AND DIDN'T HAVE 'BAM' IN IT.
    bam_classes=[]
    temp = ""
    temp = stu.stu_info_dict["BAM Supplement"]
    if (stu.stu_info_dict["Transfer Courses"] == None):
        if (temp != None):
            bam_index=0
            bam_classes = temp.split('*')
            for bam_class in range(len(bam_classes)):
                student_classes.append(bam_classes[bam_index])
                bam_index+=1

    #add classes to plos
    course_pieces={}
    r = 1
    c = 5
    index = 0
    course_piece = ""
    for course in range(len(student_classes)):
        course_pieces= student_classes[index].split(',')
        r +=1
        for p in range(len(course_pieces)):
            course_piece = course_pieces[p].strip()
            if (c == 5):
                if(len(course_piece)>8):
                    new=course_piece[0:4]
                    new = new + ' '
                    new = new + course_piece[5:]
                    course_piece = new
            cell_obj=plos_sheet.cell(r,c)
            cell_obj.value = course_piece
            c +=1
        c = 5
        index +=1
    
    #add the formulas for main MSCPS tab - TESTING IN PROGRESS
  #  if ("CSEN-MSCPS" in stu.stu_info_dict["Study Plan Codes"]):
     #   plos_sheet=plos_wb['MSCPS']
     ##   curr_cell = plos_sheet['A4']
      #  curr_cell.value= "=XLOOKUP(\"*Bin 1*\",Data!K2:K17,Data!E2:E17," ",2,1)"
      ##  curr_cell = plos_sheet['D4']
       # curr_cell.value = "=XLOOKUP(\"*Bin 2*\",Data!K2:K17,Data!E2:E17," ",2,1)"
       # curr_cell = plos_sheet['G4']
        #curr_cell.value = "=XLOOKUP(\"*Bin 3*\",Data!K2:K17,Data!E2:E17," ",2,1)"
        #curr_cell = plos_sheet['J4']
        #curr_cell.value = "=FILTER(Data!F2:H17,Data!K2:K17=\"Elective\"," ")"
        #plos_sheet['M4' ] = "=FILTER(Data!F2:H17,Data!K2:K17=\"Projects\"," ")"
        #plos_sheet['J12'] = "=SORT(FILTER(Data!F2:H17,Data!K2:K17=\"Not Applicable\"," "),1)"
        #plos_sheet['R4' ] = "=FILTER(Data!F2:H17,Data!L2:L17=E27," ")"
        #plos_sheet['U4' ] = "=FILTER(Data!F2:H17,ISNUMBER(SEARCH(\"AIG\",Data!L2:L17))," ")"
        #plos_sheet['X4' ] = "=FILTER(Data!F2:H17,ISNUMBER(SEARCH(\"ANO\",Data!L2:L17))," ")"
        #plos_sheet['AA4'] = "=FILTER(Data!F2:H17,ISNUMBER(SEARCH(\"DSE\",Data!L2:L17))," ")"
        #plos_sheet['AD4'] = "=FILTER(Data!F2:H17,ISNUMBER(SEARCH(\"HCC\",Data!L2:L17))," ")"
        #plos_sheet['AG4'] = "=FILTER(Data!F2:H17,ISNUMBER(SEARCH(\"NUM\",Data!L2:L17))," ")"
        #plos_sheet['AJ4'] = "=FILTER(Data!F2:H17,ISNUMBER(SEARCH(\"RBT\",Data!L2:L17))," ")"
        #plos_sheet['AM4'] = "=FILTER(Data!F2:H17,ISNUMBER(SEARCH(\"SEC\",Data!L2:L17))," ")"
        #plos_sheet['AP4'] = "=FILTER(Data!F2:H17,ISNUMBER(SEARCH(\"SSC\",Data!L2:L17))," ")"
        #plos_sheet['R9' ] = "=IF(E27=\"AIG\",U9:U27,IF(E27=\"ANO\",X9:X19,IF(E27=\"DSE\",AA9:AA22,IF(E27=\"HCC\",AD9:AD29,IF(E27=\"NUM\",AG9:AG17,IF(E27=\"RBT\",AJ9:AJ26,IF(E27=\"SEC\",AM9:AM23,IF(E27=\"SSC\",AP9:AP20," "))))))))"
        #plos_sheet['S9' ] = "=IF(E27=\"AIG\",INDEX(CHOOSE({1,2},V9),1),IF(E27=\"ANO\",INDEX(CHOOSE({1,2},Y9),1),IF(E27=\"DSE\",INDEX(CHOOSE({1,2},AB9),1),IF(E27=\"HCC\",INDEX(CHOOSE({1,2},AE9),1),IF(E27=\"NUM\",INDEX(CHOOSE({1,2},AH9),1),IF(E27=\"RBT\",INDEX(CHOOSE({1,2},AK9),1),IF(E27=\"SEC\",INDEX(CHOOSE({1,2},AN9),1),IF(E27=\"SSC\",INDEX(CHOOSE({1,2},AQ9),1)," "))))))))"
        #plos_sheet['S10'] = "=IF(E27=\"AIG\",INDEX(CHOOSE({1,2},V10),1),IF(E27=\"ANO\",INDEX(CHOOSE({1,2},Y10),1),IF(E27=\"DSE\",INDEX(CHOOSE({1,2},AB10),1),IF(E27=\"HCC\",INDEX(CHOOSE({1,2},AE10),1),IF(E27=\"NUM\",INDEX(CHOOSE({1,2},AH10),1),IF(E27=\"RBT\",INDEX(CHOOSE({1,2},AK10),1),IF(E27=\"SEC\",INDEX(CHOOSE({1,2},AN10),1),IF(E27=\"SSC\",INDEX(CHOOSE({1,2},AQ10),1)," "))))))))"
        #plos_sheet['S11'] = "=IF(E27=\"AIG\",INDEX(CHOOSE({1,2},V11),1),IF(E27=\"ANO\",INDEX(CHOOSE({1,2},Y11),1),IF(E27=\"DSE\",INDEX(CHOOSE({1,2},AB11),1),IF(E27=\"HCC\",INDEX(CHOOSE({1,2},AE11),1),IF(E27=\"NUM\",INDEX(CHOOSE({1,2},AH11),1),IF(E27=\"RBT\",INDEX(CHOOSE({1,2},AK11),1),IF(E27=\"SEC\",INDEX(CHOOSE({1,2},AN11),1),IF(E27=\"SSC\",INDEX(CHOOSE({1,2},AQ11),1)," "))))))))"
        #plos_sheet['S12'] = "=IF(E27=\"AIG\",INDEX(CHOOSE({1,2},V12),1),IF(E27=\"ANO\",INDEX(CHOOSE({1,2},Y12),1),IF(E27=\"DSE\",INDEX(CHOOSE({1,2},AB12),1),IF(E27=\"HCC\",INDEX(CHOOSE({1,2},AE12),1),IF(E27=\"NUM\",INDEX(CHOOSE({1,2},AH12),1),IF(E27=\"RBT\",INDEX(CHOOSE({1,2},AK12),1),IF(E27=\"SEC\",INDEX(CHOOSE({1,2},AN12),1),IF(E27=\"SSC\",INDEX(CHOOSE({1,2},AQ12),1)," "))))))))"
        #plos_sheet['S13'] = "=IF(E27=\"AIG\",INDEX(CHOOSE({1,2},V13),1),IF(E27=\"ANO\",INDEX(CHOOSE({1,2},Y13),1),IF(E27=\"DSE\",INDEX(CHOOSE({1,2},AB13),1),IF(E27=\"HCC\",INDEX(CHOOSE({1,2},AE13),1),IF(E27=\"NUM\",INDEX(CHOOSE({1,2},AH13),1),IF(E27=\"RBT\",INDEX(CHOOSE({1,2},AK13),1),IF(E27=\"SEC\",INDEX(CHOOSE({1,2},AN13),1),IF(E27=\"SSC\",INDEX(CHOOSE({1,2},AQ13),1)," "))))))))"
        #plos_sheet['S14'] = "=IF(E27=\"AIG\",INDEX(CHOOSE({1,2},V14),1),IF(E27=\"ANO\",INDEX(CHOOSE({1,2},Y14),1),IF(E27=\"DSE\",INDEX(CHOOSE({1,2},AB14),1),IF(E27=\"HCC\",INDEX(CHOOSE({1,2},AE14),1),IF(E27=\"NUM\",INDEX(CHOOSE({1,2},AH14),1),IF(E27=\"RBT\",INDEX(CHOOSE({1,2},AK14),1),IF(E27=\"SEC\",INDEX(CHOOSE({1,2},AN14),1),IF(E27=\"SSC\",INDEX(CHOOSE({1,2},AQ14),1)," "))))))))"
        #plos_sheet['S15'] = "=IF(E27=\"AIG\",INDEX(CHOOSE({1,2},V15),1),IF(E27=\"ANO\",INDEX(CHOOSE({1,2},Y15),1),IF(E27=\"DSE\",INDEX(CHOOSE({1,2},AB15),1),IF(E27=\"HCC\",INDEX(CHOOSE({1,2},AE15),1),IF(E27=\"NUM\",INDEX(CHOOSE({1,2},AH15),1),IF(E27=\"RBT\",INDEX(CHOOSE({1,2},AK15),1),IF(E27=\"SEC\",INDEX(CHOOSE({1,2},AN15),1),IF(E27=\"SSC\",INDEX(CHOOSE({1,2},AQ15),1)," "))))))))"
        #plos_sheet['S16'] = "=IF(E27=\"AIG\",INDEX(CHOOSE({1,2},V16),1),IF(E27=\"ANO\",INDEX(CHOOSE({1,2},Y16),1),IF(E27=\"DSE\",INDEX(CHOOSE({1,2},AB16),1),IF(E27=\"HCC\",INDEX(CHOOSE({1,2},AE16),1),IF(E27=\"NUM\",INDEX(CHOOSE({1,2},AH16),1),IF(E27=\"RBT\",INDEX(CHOOSE({1,2},AK16),1),IF(E27=\"SEC\",INDEX(CHOOSE({1,2},AN16),1),IF(E27=\"SSC\",INDEX(CHOOSE({1,2},AQ16),1)," "))))))))"
        #plos_sheet['S17'] = "=IF(E27=\"AIG\",INDEX(CHOOSE({1,2},V17),1),IF(E27=\"ANO\",INDEX(CHOOSE({1,2},Y17),1),IF(E27=\"DSE\",INDEX(CHOOSE({1,2},AB17),1),IF(E27=\"HCC\",INDEX(CHOOSE({1,2},AE17),1),IF(E27=\"NUM\",INDEX(CHOOSE({1,2},AH17),1),IF(E27=\"RBT\",INDEX(CHOOSE({1,2},AK17),1),IF(E27=\"SEC\",INDEX(CHOOSE({1,2},AN17),1),IF(E27=\"SSC\",INDEX(CHOOSE({1,2},AQ17),1)," "))))))))"
        #plos_sheet['S18'] = "=IF(E27=\"AIG\",INDEX(CHOOSE({1,2},V18),1),IF(E27=\"ANO\",INDEX(CHOOSE({1,2},Y18),1),IF(E27=\"HCC\",INDEX(CHOOSE({1,2},AE18),1),IF(E27=\"DSE\",INDEX(CHOOSE({1,2},AB18),1),IF(E27=\"RBT\",INDEX(CHOOSE({1,2},AK18),1),IF(E27=\"SEC\",INDEX(CHOOSE({1,2},AN18),1),IF(E27=\"SSC\",INDEX(CHOOSE({1,2},AQ18),1)," ")))))))"
        #plos_sheet['S19'] = "=IF(E27=\"AIG\",INDEX(CHOOSE({1,2},V19),1),IF(E27=\"ANO\",INDEX(CHOOSE({1,2},Y19),1),IF(E27=\"HCC\",INDEX(CHOOSE({1,2},AE19),1),IF(E27=\"DSE\",INDEX(CHOOSE({1,2},AB19),1),IF(E27=\"RBT\",INDEX(CHOOSE({1,2},AK19),1),IF(E27=\"SEC\",INDEX(CHOOSE({1,2},AN19),1),IF(E27=\"SSC\",INDEX(CHOOSE({1,2},AQ19),1)," ")))))))"
        #plos_sheet['S20'] = "=IF(E27=\"AIG\",INDEX(CHOOSE({1,2},V20),1),IF(E27=\"HCC\",INDEX(CHOOSE({1,2},AE20),1),IF(E27=\"DSE\",INDEX(CHOOSE({1,2},AB20),1),IF(E27=\"RBT\",INDEX(CHOOSE({1,2},AK20),1),IF(E27=\"SEC\",INDEX(CHOOSE({1,2},AN20),1),IF(E27=\"SSC\",INDEX(CHOOSE({1,2},AQ20),1)," "))))))"
        #plos_sheet['S21'] = "=IF(E27=\"AIG\",INDEX(CHOOSE({1,2},V21),1),IF(E27=\"HCC\",INDEX(CHOOSE({1,2},AE21),1),IF(E27=\"DSE\",INDEX(CHOOSE({1,2},AB21),1),IF(E27=\"RBT\",INDEX(CHOOSE({1,2},AK21),1),IF(E27=\"SEC\",INDEX(CHOOSE({1,2},AN21),1)," ")))))"
        #plos_sheet['S22'] = "=IF(E27=\"AIG\",INDEX(CHOOSE({1,2},V23),1),IF(E27=\"HCC\",INDEX(CHOOSE({1,2},AE23),1),IF(E27=\"RBT\",INDEX(CHOOSE({1,2},AK23),1),IF(E27=\SEC\",INDEX(CHOOSE({1,2},AN23),1)," "))))"
        #plos_sheet['S23'] = "=IF(E27=\"AIG\",INDEX(CHOOSE({1,2},V23),1),IF(E27=\"HCC\",INDEX(CHOOSE({1,2},AE23),1),IF(E27=\"RBT\",INDEX(CHOOSE({1,2},AK23),1),IF(E27=\"SEC\",INDEX(CHOOSE({1,2},AN23),1)," "))))"
        #plos_sheet['S24'] = "=IF(E27=\"AIG\",INDEX(CHOOSE({1,2},V24),1),IF(E27=\"HCC\",INDEX(CHOOSE({1,2},AE24),1),IF(E27=\"RBT\",INDEX(CHOOSE({1,2},AK24),1)," ")))"
        #plos_sheet['S25'] = "=IF(E27=\"AIG\",INDEX(CHOOSE({1,2},V25),1),IF(E27=\"HCC\",INDEX(CHOOSE({1,2},AE27),1),IF(E27=\"RBT\",INDEX(CHOOSE({1,2},AK25),1)," ")))"
        #plos_sheet['S26'] = "=IF(E27=\"AIG\",INDEX(CHOOSE({1,2},V26),1),IF(E27=\"HCC\",INDEX(CHOOSE({1,2},AE26),1),IF(E27=\"RBT\",INDEX(CHOOSE({1,2},AK26),1)," ")))"
        #plos_sheet['S27'] = "=IF(E27=\"AIG\",INDEX(CHOOSE({1,2},V27),1),IF(E27=\"HCC\",INDEX(CHOOSE({1,2},AE27),1)," "))"
        #plos_sheet['S28'] = "IF(E27=\"HCC\",INDEX(CHOOSE({1,2},AE28),1)," ")"
        #plos_sheet['S29'] = "IF(E27=\"HCC\",INDEX(CHOOSE({1,2},AE29),1)," ")"
   
    #add the formulas for main MSNE tab
    #elif ("NTEN-MSNE" in stu.stu_info_dict["Study Plan Codes"]):
        #plos_sheet=plos_wb['MSNE']
        

    #close out
    plos_wb.save(new_file_path)
    student_classes.clear()
    bam_classes.clear()
    transfer_classes.clear()
    course_pieces.clear()
