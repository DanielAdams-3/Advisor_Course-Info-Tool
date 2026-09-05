import pandas as pd
import json
import js2py
import requests
import openpyxl
import re #copilot

js_file = "courseCatalog.js"
js2_file = "topicsCourseCatalog.js"

context=js2py.EvalJs()

def clean(value):
    """
    Convert to string,
    remove trailing asterisks,
    remove extra whitespace.
    """
    return str(value).strip()

with open("courseCatalog.js", "r", encoding="utf-8") as g:
    catalog_js = g.read()

    #error, function is not accurately 
    context.execute(catalog_js)
    context.execute(""" 
        function retrieveCourseObject(requestedSubject) { 
            for (let i=0; i< courseCatalog.length;i++)
            {
                if (requestedSubject === undefined)
                {
                    return null;
                }
                currCourseSubj=courseCatalog[i]['courseSubject'];
                if (requestedSubject === currCourseSubj)
                {
                let result=courseCatalog[i];
                    return result;
                }
            }
            return null;
        }
    """)

#Step 1 - open the google sheet
#https://stackoverflow.com/questions/64746066/how-to-access-google-sheets-without-authentication
#https://medium.com/@Bwhiz/step-by-step-guide-importing-google-sheets-data-into-pandas-ae2df899257f

html = '<a href="https://docs.google.com/spreadsheets/d/11YBH99TtZ-5n22z8hDVdbpa43s2aK7L7crpLpRsKXeo/export?format=xlsx">link</a>'
match = re.search(r'href="([^"]+)"', html)
if match:
    xlsx_url = match.group(1)
response = requests.get(xlsx_url)
response.raise_for_status()


with open("original.xlsx", "wb") as f:
    f.write(response.content)
f.close()

wb=openpyxl.load_workbook("original.xlsx")
active_sheet=wb.active

c=1
row_num = 3
cell_obj=active_sheet.cell(row_num,c)
offerings=""
csphd=""
csms=""
mscps=""
msne=""
msai=""
notes=""

while (cell_obj.value != "end"):

    #FIXME: need to account for CSCI 7412, 7712, etc.

    requestedSubject=clean(cell_obj.value)
    result=context.retrieveCourseObject(requestedSubject) 
    if ((result is None) or (type(result) == "NoneType")):
        offerings=""
        csphd=""
        csms=""
        mscps=""
        msne=""
        msai=""
        notes=""
    else:
        offerings=result.offerings        
        csphd=result.reqCSENPHD     
        csms=result.reqCSENMS       
        mscps=result.reqCSENMSCPS
        msne=result.reqNTENMSNE     
        msai=result.reqAINTMSAI     
        notes=result.reqNote     

    #assign values to correct cells
    c=3
    
    cell_obj=active_sheet.cell(row_num,c)
    cell_obj.value=notes        #col3
    c=5

    cell_obj=active_sheet.cell(row_num,c)
    cell_obj.value=offerings    #col 5
    c+=1

    cell_obj=active_sheet.cell(row_num,c)
    cell_obj.value=csphd        #col 6
    c+=1      

    cell_obj=active_sheet.cell(row_num,c)
    cell_obj.value=csms         #col 7
    c+=1

    cell_obj=active_sheet.cell(row_num,c)
    cell_obj.value=mscps        #col 8
    c+=1

    cell_obj=active_sheet.cell(row_num,c)
    cell_obj.value=msne         #col 9
    c+=1

    cell_obj=active_sheet.cell(row_num,c)
    cell_obj.value=msai         #col10

    #go to next row
    row_num += 1
    c=1
    cell_obj=active_sheet.cell(row_num,c)
    wb.save("updated.xlsx")



#Process Topics Courses
#chagne the active sheet
with open("topicsCourseCatalog.js", "r", encoding="utf-8") as h:
    topicsCatalog_js = h.read()

    context.execute(topicsCatalog_js)
    context.execute(""" 
        function retrieveTopicsCourseObject(requestedSubject, requestedTopic) { 
            lookupSubject=clean(requestedSubject.subtring(0,9));
            lookupTopic=clean(requestedTopic.substring(0,7));
            for (let i=0; i< topicsCourseCatalog.length;i++)
            {
                if ((lookupSubject === undefined) || (lookupTitle === undefined))
                {
                    return null;
                }
                currCourseTitle=topicsCourseCatalog[i]['title'];
                currCourseSubj=topicsCourseCatalog[i]['courseSubject'];
                
                if ((lookupSubject === currCourseSubj) && (currCourseTitle.includes(lookupTitle)))
                {
                    let result=topicsCourseCatalog[i];
                    return result;
                }
            }
            return null;
        }
    """)

#switch tab to topics
active_sheet=wb['topics']
while (cell_obj.value != "end"):
    requestedSubject=clean(cell_obj.value)
    result=context.retrieveTopicsCourseObject(requestedSubject) 
    if ((result is None) or (type(result) == "NoneType")):
        offerings=""
        csphd=""
        csms=""
        mscps=""
        msne=""
        msai=""
        notes=""
    else:
        offerings=result.offerings        
        csphd=result.reqCSENPHD     
        csms=result.reqCSENMS       
        mscps=result.reqCSENMSCPS
        msne=result.reqNTENMSNE     
        msai=result.reqAINTMSAI     
        notes=result.reqNote     

    #assign values to correct cells
    c=3
    
    cell_obj=active_sheet.cell(row_num,c)
    cell_obj.value=notes        #col3
    c=5

    cell_obj=active_sheet.cell(row_num,c)
    cell_obj.value=offerings    #col 5
    c+=1

    cell_obj=active_sheet.cell(row_num,c)
    cell_obj.value=csphd        #col 6
    c+=1      

    cell_obj=active_sheet.cell(row_num,c)
    cell_obj.value=csms         #col 7
    c+=1

    cell_obj=active_sheet.cell(row_num,c)
    cell_obj.value=mscps        #col 8
    c+=1

    cell_obj=active_sheet.cell(row_num,c)
    cell_obj.value=msne         #col 9
    c+=1

    cell_obj=active_sheet.cell(row_num,c)
    cell_obj.value=msai         #col10

    #go to next row
    row_num += 1
    c=1
    cell_obj=active_sheet.cell(row_num,c)
    wb.save("updated.xlsx")

g.close()
h.close()
f.close()